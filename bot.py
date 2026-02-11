"""
CRM-Помощник — Telegram-бот поддержки отдела 1С CRM.
Сбор ошибок и предложений с записью в Excel.
"""

import json
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from telegram import (
    BotCommand,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    Update,
)
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

from config import (
    ADMIN_IDS,
    BOT_TOKEN,
    ERROR_CATEGORIES,
    EXCEL_FILE,
    GROUP_CHAT_ID,
    MODULES,
    USERS_DB_FILE,
)

# ── Состояния ConversationHandler ──────────────────────────────────────
(
    REG_FIO,
    REG_MODULE,
    MAIN_MENU,
    ERROR_CATEGORY,
    ERROR_DESCRIPTION,
    SUGGESTION_TEXT,
) = range(6)

# ── Иконки для модулей ────────────────────────────────────────────────
MODULE_EMOJI = {
    "Модуль экономической эффективности и аналитики": "📊",
    "Модуль развития цепей поставок и складской логистики": "🚛",
    "Модуль развития бизнеса 1": "💼",
    "Модуль развития бизнеса 2": "📈",
    "Модуль технологии и эффективности": "⚙️",
}

# ── Иконки для категорий ошибок ───────────────────────────────────────
ERROR_EMOJI = {
    "Воронка продаж": "📊",
    "Проблема с карточкой клиента": "👤",
    "Проблема с карточкой интереса": "📋",
    "Другое": "🔧",
}

# ── Вспомогательные функции для хранения пользователей ─────────────────

def _ensure_data_dir():
    Path(USERS_DB_FILE).parent.mkdir(parents=True, exist_ok=True)
    Path(EXCEL_FILE).parent.mkdir(parents=True, exist_ok=True)


def _load_users() -> dict:
    _ensure_data_dir()
    if os.path.exists(USERS_DB_FILE):
        with open(USERS_DB_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_users(users: dict):
    _ensure_data_dir()
    with open(USERS_DB_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)


def _get_user(user_id: int) -> dict | None:
    users = _load_users()
    return users.get(str(user_id))


def _save_user(user_id: int, fio: str, module: str):
    users = _load_users()
    users[str(user_id)] = {
        "fio": fio,
        "module": module,
        "registered_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    _save_users(users)


# ── Excel ──────────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    "Дата и время",
    "Telegram ID",
    "ФИО",
    "Модуль",
    "Тип обращения",
    "Категория ошибки",
    "Описание",
]


def _ensure_excel():
    _ensure_data_dir()
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Обращения"
        ws.append(EXCEL_HEADERS)
        widths = [20, 14, 25, 22, 20, 30, 60]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        wb.save(EXCEL_FILE)


def _append_to_excel(row: list):
    _ensure_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(row)
    wb.save(EXCEL_FILE)


# ── Хранение заявок (для отслеживания «Взять в работу») ────────────────

TICKETS_FILE = "data/tickets.json"


def _load_tickets() -> dict:
    _ensure_data_dir()
    if os.path.exists(TICKETS_FILE):
        with open(TICKETS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"next_id": 1, "items": {}}


def _save_tickets(tickets: dict):
    _ensure_data_dir()
    with open(TICKETS_FILE, "w", encoding="utf-8") as f:
        json.dump(tickets, f, ensure_ascii=False, indent=2)


def _create_ticket(ticket_type: str, user_fio: str, module: str,
                   category: str, description: str) -> int:
    """Создаёт заявку и возвращает её ID."""
    tickets = _load_tickets()
    tid = tickets["next_id"]
    tickets["next_id"] = tid + 1
    tickets["items"][str(tid)] = {
        "type": ticket_type,
        "fio": user_fio,
        "module": module,
        "category": category,
        "description": description,
        "status": "new",
        "taken_by": None,
        "admin_messages": {},
        "group_message_id": None,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    _save_tickets(tickets)
    return tid


def _take_ticket(tid: int, admin_name: str) -> dict | None:
    """Помечает заявку как взятую. Возвращает данные заявки."""
    tickets = _load_tickets()
    ticket = tickets["items"].get(str(tid))
    if not ticket:
        return None
    ticket["status"] = "in_progress"
    ticket["taken_by"] = admin_name
    _save_tickets(tickets)
    return ticket


def _save_ticket_message(tid: int, admin_id: int, message_id: int):
    """Сохраняет message_id уведомления для конкретного админа."""
    tickets = _load_tickets()
    ticket = tickets["items"].get(str(tid))
    if ticket:
        ticket["admin_messages"][str(admin_id)] = message_id
        _save_tickets(tickets)


def _save_ticket_group_message(tid: int, message_id: int):
    """Сохраняет message_id уведомления в группе."""
    tickets = _load_tickets()
    ticket = tickets["items"].get(str(tid))
    if ticket:
        ticket["group_message_id"] = message_id
        _save_tickets(tickets)


# ── Клавиатуры ─────────────────────────────────────────────────────────

def _main_menu_keyboard() -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton("🐞 Сообщить об ошибке", callback_data="report_error")],
        [InlineKeyboardButton("💡 Предложить улучшение", callback_data="suggest")],
    ]
    return InlineKeyboardMarkup(buttons)


def _modules_keyboard() -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(
            f"{MODULE_EMOJI.get(m, '📁')} {m}",
            callback_data=f"module:{i}",
        )]
        for i, m in enumerate(MODULES)
    ]
    return InlineKeyboardMarkup(buttons)


def _error_categories_keyboard() -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(
            f"{ERROR_EMOJI.get(c, '❓')} {c}",
            callback_data=f"errcat:{c}",
        )]
        for c in ERROR_CATEGORIES
    ]
    buttons.append(
        [InlineKeyboardButton("« Назад", callback_data="back_menu")]
    )
    return InlineKeyboardMarkup(buttons)


def _cancel_keyboard() -> InlineKeyboardMarkup:
    """Кнопка отмены на этапах ввода текста."""
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton("✕ Отмена", callback_data="back_menu")]]
    )


def _back_to_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton("« В главное меню", callback_data="back_menu")]]
    )


# Постоянная клавиатура внизу экрана
PERSISTENT_KEYBOARD = ReplyKeyboardMarkup(
    [[KeyboardButton("▶️ Старт")]],
    resize_keyboard=True,
)


# ── Хендлеры ───────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Точка входа: /start. Проверяем регистрацию."""
    user_id = update.effective_user.id
    user = _get_user(user_id)

    if user:
        # Если модуль устарел — просим выбрать заново
        if user["module"] not in MODULES:
            context.user_data["reg_fio"] = user["fio"]
            await update.message.reply_text(
                f"Здравствуйте, <b>{user['fio']}</b>!\n\n"
                "Список модулей обновился.\n"
                "Пожалуйста, выберите ваш модуль заново:",
                reply_markup=_modules_keyboard(),
                parse_mode="HTML",
            )
            return REG_MODULE
        return await _show_main_menu(update, context, user)

    await update.message.reply_text(
        "Добро пожаловать в <b>CRM-Помощник</b>! 👋\n\n"
        "Здесь вы можете сообщить об ошибке "
        "или предложить улучшение для 1С CRM.\n\n"
        "Для начала давайте познакомимся.\n"
        "Введите ваше ФИО:",
        reply_markup=PERSISTENT_KEYBOARD,
        parse_mode="HTML",
    )
    return REG_FIO


async def reg_fio(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Получаем ФИО, предлагаем выбрать модуль."""
    fio = update.message.text.strip()
    if len(fio) < 3:
        await update.message.reply_text("Пожалуйста, введите корректное ФИО:")
        return REG_FIO

    context.user_data["reg_fio"] = fio
    await update.message.reply_text(
        f"Отлично, <b>{fio}</b>!\n\n"
        "Выберите модуль 1С CRM, с которым вы работаете:",
        reply_markup=_modules_keyboard(),
        parse_mode="HTML",
    )
    return REG_MODULE


async def reg_module(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Сохраняем модуль, показываем главное меню."""
    query = update.callback_query
    await query.answer()

    module_index = int(query.data.removeprefix("module:"))
    module = MODULES[module_index]
    fio = context.user_data.pop("reg_fio")
    user_id = update.effective_user.id

    _save_user(user_id, fio, module)

    user = _get_user(user_id)
    return await _show_main_menu_from_callback(query, context, user)


async def _show_main_menu(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    user: dict,
) -> int:
    """Главное меню (из обычного сообщения)."""
    emoji = MODULE_EMOJI.get(user["module"], "📁")
    await update.message.reply_text(
        f"Здравствуйте, <b>{user['fio']}</b>!\n"
        f"{emoji} Модуль: {user['module']}\n\n"
        "Выберите действие:",
        reply_markup=_main_menu_keyboard(),
        parse_mode="HTML",
    )
    return MAIN_MENU


async def _show_main_menu_from_callback(query, context, user: dict) -> int:
    """Главное меню (из callback-кнопки)."""
    emoji = MODULE_EMOJI.get(user["module"], "📁")
    await query.edit_message_text(
        f"Здравствуйте, <b>{user['fio']}</b>!\n"
        f"{emoji} Модуль: {user['module']}\n\n"
        "Выберите действие:",
        reply_markup=_main_menu_keyboard(),
        parse_mode="HTML",
    )
    return MAIN_MENU


async def menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка нажатий главного меню."""
    query = update.callback_query
    await query.answer()

    if query.data == "report_error":
        await query.edit_message_text(
            "🐞 <b>Сообщить об ошибке</b>\n\n"
            "Выберите категорию проблемы:",
            reply_markup=_error_categories_keyboard(),
            parse_mode="HTML",
        )
        return ERROR_CATEGORY

    if query.data == "suggest":
        await query.edit_message_text(
            "💡 <b>Предложить улучшение</b>\n\n"
            "Опишите, что можно улучшить в системе.\n"
            "Любая деталь может быть полезной.",
            reply_markup=_cancel_keyboard(),
            parse_mode="HTML",
        )
        return SUGGESTION_TEXT

    return MAIN_MENU


async def error_category_handler(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    """Выбор категории ошибки."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_menu":
        user = _get_user(update.effective_user.id)
        return await _show_main_menu_from_callback(query, context, user)

    category = query.data.removeprefix("errcat:")
    context.user_data["error_category"] = category

    cat_emoji = ERROR_EMOJI.get(category, "🔧")

    if category == "Другое":
        await query.edit_message_text(
            f"{cat_emoji} <b>Категория: Другое</b>\n\n"
            "Расскажите подробнее, с какой проблемой вы столкнулись.\n"
            "Опишите шаги, которые привели к ошибке — "
            "это поможет нам разобраться быстрее.",
            reply_markup=_cancel_keyboard(),
            parse_mode="HTML",
        )
        return ERROR_DESCRIPTION

    await query.edit_message_text(
        f"{cat_emoji} <b>Категория: {category}</b>\n\n"
        "Опишите проблему:\n"
        "• Что произошло?\n"
        "• При каких действиях?\n"
        "• Есть ли скриншот?",
        reply_markup=_cancel_keyboard(),
        parse_mode="HTML",
    )
    return ERROR_DESCRIPTION


async def error_description_handler(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    """Получаем описание ошибки, сохраняем в Excel."""
    user_id = update.effective_user.id
    user = _get_user(user_id)
    category = context.user_data.pop("error_category", "—")
    description = update.message.text.strip()

    _append_to_excel([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        user_id,
        user["fio"],
        user["module"],
        "Ошибка",
        category,
        description,
    ])

    await update.message.reply_text(
        "✅ <b>Принято в работу!</b>\n\n"
        "Спасибо, что сообщили — мы разберёмся "
        "и постараемся исправить.",
        reply_markup=_back_to_menu_keyboard(),
        parse_mode="HTML",
    )

    # Создаём заявку
    tid = _create_ticket("Ошибка", user["fio"], user["module"],
                         category, description)

    # Уведомление админам и в группу
    await _notify_new_ticket(context, tid, "🚨", "Ошибка",
                             user["fio"], user["module"],
                             category, description)

    return MAIN_MENU


async def suggestion_text_handler(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    """Получаем предложение, сохраняем в Excel."""
    user_id = update.effective_user.id
    user = _get_user(user_id)
    description = update.message.text.strip()

    _append_to_excel([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        user_id,
        user["fio"],
        user["module"],
        "Предложение",
        "—",
        description,
    ])

    await update.message.reply_text(
        "✅ <b>Предложение принято!</b>\n\n"
        "Спасибо за идею — мы обязательно рассмотрим.",
        reply_markup=_back_to_menu_keyboard(),
        parse_mode="HTML",
    )

    # Создаём заявку
    tid = _create_ticket("Предложение", user["fio"], user["module"],
                         "—", description)

    # Уведомление админам и в группу
    await _notify_new_ticket(context, tid, "💡", "Предложение",
                             user["fio"], user["module"],
                             "—", description)

    return MAIN_MENU


async def back_to_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Возврат в главное меню из callback."""
    query = update.callback_query
    await query.answer()
    user = _get_user(update.effective_user.id)
    if user:
        return await _show_main_menu_from_callback(query, context, user)
    await query.edit_message_text("Нажмите /start для начала.")
    return ConversationHandler.END


# ── Уведомления по заявкам ─────────────────────────────────────────────

def _ticket_text(emoji: str, ticket_type: str, fio: str, module: str,
                 category: str, description: str, tid: int) -> str:
    """Формирует текст уведомления о заявке."""
    lines = [
        f"{emoji} <b>Новая заявка #{tid}: {ticket_type}</b>\n",
        f"👤 {fio}",
        f"📦 {module}",
    ]
    if category and category != "—":
        lines.append(f"📂 {category}")
    lines.append(f"💬 {description}")
    return "\n".join(lines)


async def _notify_new_ticket(context, tid: int, emoji: str, ticket_type: str,
                             fio: str, module: str, category: str,
                             description: str):
    """Отправляет уведомление админам и в группу с кнопкой «Взять в работу»."""
    text = _ticket_text(emoji, ticket_type, fio, module,
                        category, description, tid)
    keyboard = InlineKeyboardMarkup(
        [[InlineKeyboardButton("🙋 Взять в работу", callback_data=f"take:{tid}")]]
    )

    # Личные сообщения админам
    for admin_id in ADMIN_IDS:
        try:
            msg = await context.bot.send_message(
                admin_id, text, parse_mode="HTML", reply_markup=keyboard,
            )
            _save_ticket_message(tid, admin_id, msg.message_id)
        except Exception:
            pass

    # Сообщение в группу
    if GROUP_CHAT_ID:
        try:
            msg = await context.bot.send_message(
                GROUP_CHAT_ID, text, parse_mode="HTML", reply_markup=keyboard,
            )
            _save_ticket_group_message(tid, msg.message_id)
        except Exception:
            pass


async def take_ticket_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Админ нажал «Взять в работу»."""
    query = update.callback_query
    user_id = update.effective_user.id

    if user_id not in ADMIN_IDS:
        await query.answer("У вас нет доступа.", show_alert=True)
        return

    tid = int(query.data.removeprefix("take:"))
    tickets = _load_tickets()
    ticket = tickets["items"].get(str(tid))

    if not ticket:
        await query.answer("Заявка не найдена.", show_alert=True)
        return

    if ticket["status"] != "new":
        await query.answer(
            f"Уже взята: {ticket['taken_by']}", show_alert=True,
        )
        return

    admin_name = update.effective_user.full_name
    ticket = _take_ticket(tid, admin_name)
    await query.answer("Вы взяли заявку в работу!")

    # Обновлённый текст
    emoji = "🚨" if ticket["type"] == "Ошибка" else "💡"
    base_text = _ticket_text(emoji, ticket["type"], ticket["fio"],
                             ticket["module"], ticket["category"],
                             ticket["description"], tid)
    updated_text = base_text + f"\n\n✅ <b>Взял(а): {admin_name}</b>"

    # Обновить сообщения у всех админов
    for aid_str, mid in ticket["admin_messages"].items():
        try:
            await context.bot.edit_message_text(
                updated_text, chat_id=int(aid_str), message_id=mid,
                parse_mode="HTML",
            )
        except Exception:
            pass

    # Обновить сообщение в группе
    if GROUP_CHAT_ID and ticket["group_message_id"]:
        try:
            await context.bot.edit_message_text(
                updated_text, chat_id=GROUP_CHAT_ID,
                message_id=ticket["group_message_id"],
                parse_mode="HTML",
            )
        except Exception:
            pass


# ── Админские команды ──────────────────────────────────────────────────

async def cmd_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показать админ-панель."""
    user_id = update.effective_user.id
    if user_id not in ADMIN_IDS:
        await update.message.reply_text("У вас нет доступа к этой команде.")
        return

    buttons = [
        [InlineKeyboardButton("📥 Выгрузить Excel", callback_data="admin:export")],
        [InlineKeyboardButton("📊 Статистика", callback_data="admin:stats")],
        [InlineKeyboardButton("👥 Список пользователей", callback_data="admin:users")],
    ]
    await update.message.reply_text(
        "⚙️ <b>Панель администратора</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML",
    )


async def admin_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка админских кнопок."""
    query = update.callback_query
    await query.answer()

    user_id = update.effective_user.id
    if user_id not in ADMIN_IDS:
        await query.edit_message_text("У вас нет доступа.")
        return

    action = query.data.removeprefix("admin:")

    if action == "export":
        _ensure_excel()
        if os.path.exists(EXCEL_FILE):
            await query.message.reply_document(
                document=open(EXCEL_FILE, "rb"),
                filename=f"crm_support_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                caption="Выгрузка обращений",
            )
        else:
            await query.edit_message_text("Файл обращений пока пуст.")

    elif action == "stats":
        _ensure_excel()
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            total = ws.max_row - 1
            errors = sum(1 for row in ws.iter_rows(min_row=2) if row[4].value == "Ошибка")
            suggestions = sum(1 for row in ws.iter_rows(min_row=2) if row[4].value == "Предложение")
            users = _load_users()
            text = (
                "📊 <b>Статистика</b>\n\n"
                f"Всего обращений: <b>{total}</b>\n"
                f"Ошибок: <b>{errors}</b>\n"
                f"Предложений: <b>{suggestions}</b>\n"
                f"Пользователей: <b>{len(users)}</b>"
            )
        else:
            text = "Данных пока нет."
        await query.edit_message_text(text, parse_mode="HTML")

    elif action == "users":
        users = _load_users()
        if not users:
            await query.edit_message_text("Зарегистрированных пользователей нет.")
            return
        lines = []
        for uid, info in users.items():
            m_emoji = MODULE_EMOJI.get(info["module"], "📁")
            lines.append(f"{info['fio']} — {m_emoji} {info['module']} (ID: <code>{uid}</code>)")
        text = "👥 <b>Пользователи</b>\n\n" + "\n".join(lines)
        if len(text) > 4000:
            text = text[:4000] + "\n\n... (список обрезан)"
        await query.edit_message_text(text, parse_mode="HTML")


# ── Настройка команд бота (кнопка «Меню» в Telegram) ──────────────────

async def post_init(application):
    """Устанавливаем команды бота — они появятся в кнопке Меню."""
    await application.bot.set_my_commands([
        BotCommand("start", "Главное меню"),
        BotCommand("admin", "Панель администратора"),
    ])


async def on_added_to_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Когда бота добавляют в группу — показать chat_id."""
    chat = update.effective_chat
    if chat.type not in ("group", "supergroup"):
        return
    bot_user = await context.bot.get_me()
    for member in update.message.new_chat_members:
        if member.id == bot_user.id:
            await update.message.reply_text(
                f"👋 Бот добавлен в группу!\n\n"
                f"<b>Chat ID этой группы:</b>\n"
                f"<code>{chat.id}</code>\n\n"
                f"Скопируйте и вставьте в <code>config.py</code>:\n"
                f"<code>GROUP_CHAT_ID = {chat.id}</code>",
                parse_mode="HTML",
            )
            break


# ── Запуск ─────────────────────────────────────────────────────────────

def main():
    _ensure_data_dir()

    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()

    # Обработка текстовой кнопки «▶️ Старт»
    async def text_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        return await cmd_start(update, context)

    # Основной диалог
    conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", cmd_start),
            MessageHandler(filters.Regex(r"^▶️ Старт$"), text_start),
        ],
        states={
            REG_FIO: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, reg_fio),
            ],
            REG_MODULE: [
                CallbackQueryHandler(reg_module, pattern=r"^module:"),
            ],
            MAIN_MENU: [
                CallbackQueryHandler(menu_handler, pattern=r"^(report_error|suggest)$"),
                CallbackQueryHandler(back_to_menu, pattern=r"^back_menu$"),
            ],
            ERROR_CATEGORY: [
                CallbackQueryHandler(error_category_handler, pattern=r"^errcat:"),
                CallbackQueryHandler(back_to_menu, pattern=r"^back_menu$"),
            ],
            ERROR_DESCRIPTION: [
                CallbackQueryHandler(back_to_menu, pattern=r"^back_menu$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, error_description_handler),
            ],
            SUGGESTION_TEXT: [
                CallbackQueryHandler(back_to_menu, pattern=r"^back_menu$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, suggestion_text_handler),
            ],
        },
        fallbacks=[
            CommandHandler("start", cmd_start),
            MessageHandler(filters.Regex(r"^▶️ Старт$"), text_start),
        ],
    )

    app.add_handler(conv)

    # Админка (вне ConversationHandler, чтобы работала всегда)
    app.add_handler(CommandHandler("admin", cmd_admin))
    app.add_handler(CallbackQueryHandler(admin_callback, pattern=r"^admin:"))
    app.add_handler(CallbackQueryHandler(take_ticket_callback, pattern=r"^take:"))

    # Подсказка chat_id при добавлении бота в группу
    app.add_handler(MessageHandler(filters.StatusUpdate.NEW_CHAT_MEMBERS, on_added_to_group))

    print("CRM-Помощник запущен...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
